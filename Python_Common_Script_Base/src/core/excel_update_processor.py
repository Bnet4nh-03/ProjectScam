import openpyxl


class ExcelUpdateProcessor:
    def __init__(self, config, logger):
        self.logger = logger
        self.database = config["database"]
        self.table_name = config["table_name"]
        self.mapping = config["header_mapping"]
        self.key_headers = set(config["key_columns"])

    def _is_colored(self, cell):
        fill = cell.fill
        if fill is None or fill.start_color is None:
            return False

        rgb = fill.start_color.rgb

        if rgb is None or rgb == "00000000":
            return False

        return True

    def process_file(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # map header
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value in self.mapping:
                headers[col_idx] = {
                    "excel": cell.value,
                    "db": self.mapping[cell.value]
                }

        payloads = []

        # loop row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            keys = {}
            update_fields = {}

            for col_idx, cell in enumerate(row, start=1):

                if col_idx not in headers:
                    continue

                header_info = headers[col_idx]
                excel_header = header_info["excel"]
                db_column = header_info["db"]

                value = cell.value

                # ✅ 1. KEY → luôn lấy
                if excel_header in self.key_headers:
                    if value is not None:
                        keys[db_column] = value
                    continue

                # ✅ 2. NON-KEY → check màu
                if not self._is_colored(cell):
                    continue

                update_fields[db_column] = value

            # ✅ skip nếu không có update
            if not update_fields:
                continue

            # ✅ skip nếu thiếu key
            if len(keys) != len(self.key_headers):
                self.logger.warning(f"Row {row_idx} missing key → skip")
                continue

            payloads.append({
                "database": self.database,
                "table": self.table_name,
                "keys": keys,
                "data": update_fields
            })

        self.logger.info(f"Generated {len(payloads)} payloads")
        return payloads