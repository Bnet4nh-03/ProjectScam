import re
import logging
from rapidfuzz import fuzz
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)


class ExcelDBMapper:

    def __init__(self, db_columns):
        self.db_columns = db_columns
        self.db_norm = {col: self._normalize(col) for col in db_columns}

    # =========================
    # ✅ Normalize
    # =========================
    def _normalize(self, text):
        text = str(text).lower()
        text = re.sub(r'[^a-z0-9]', '', text)
        return text

    # =========================
    # ✅ Token split
    # =========================
    def _tokens(self, text):
        return [t for t in re.split(r'[\s_]+', text.lower()) if t]

    # =========================
    # ✅ Token similarity (Jaccard)
    # =========================
    def _token_score(self, a, b):
        a_set = set(self._tokens(a))
        b_set = set(self._tokens(b))
        return len(a_set & b_set) / max(len(a_set | b_set), 1)

    # =========================
    # ✅ Ngram similarity
    # =========================
    def _ngram(self, s, n=3):
        return [s[i:i+n] for i in range(len(s) - n + 1)] if len(s) >= n else [s]

    def _ngram_score(self, a, b):
        a_set = set(self._ngram(a))
        b_set = set(self._ngram(b))
        return len(a_set & b_set) / max(len(b_set), 1)

    # =========================
    # ✅ FINAL SCORE
    # =========================
    def _score(self, excel_col, db_col):
        norm_e = self._normalize(excel_col)
        norm_d = self._normalize(db_col)

        token = self._token_score(excel_col, db_col)
        ngram = self._ngram_score(norm_e, norm_d)
        fuzzy_score = fuzz.partial_ratio(norm_e, norm_d) / 100

        score = (
            0.45 * token +
            0.35 * ngram +
            0.20 * fuzzy_score
        )

        return score

    # =========================
    # ✅ Read Excel header + detect color
    # =========================
    def get_excel_headers_with_color(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        headers = []
        colored_headers = []

        for col in ws.iter_cols(min_row=1, max_row=1):
            cell = col[0]
            header = str(cell.value).strip() if cell.value else ""
            headers.append(header)

            fill = cell.fill
            color = fill.start_color

            # ✅ Detect column có màu (không phải trắng)
            if color and color.index not in ("00000000", "FFFFFFFF"):
                colored_headers.append(header)

        return {
            "all_headers": headers,
            "update_headers": colored_headers
        }

    # =========================
    # ✅ Mapping Engine (NO DUPLICATE)
    # =========================
    def map(self, excel_headers, key_columns=None):
        key_columns = key_columns or []

        mapped = {}
        unmapped = []
        ambiguous = []
        used_db = set()

        # =====================
        # ✅ STEP 1: KEY STRICT
        # =====================
        for key in key_columns:
            found = None
            for db_col in self.db_columns:
                if self._normalize(key) == self._normalize(db_col):
                    found = db_col
                    break

            if not found:
                raise Exception(f"Key column '{key}' không tồn tại trong DB")

            mapped[key] = found
            used_db.add(found)

        # =====================
        # ✅ STEP 2: SCORE MATCH
        # =====================
        for excel_col in excel_headers:
            if excel_col in mapped:
                continue

            best_col = None
            best_score = 0

            for db_col in self.db_columns:
                if db_col in used_db:
                    continue  # ✅ tránh duplicate

                score = self._score(excel_col, db_col)

                if score > best_score:
                    best_score = score
                    best_col = db_col

            # =====================
            # ✅ DECISION
            # =====================
        
            mapped[excel_col] = best_col
            used_db.add(best_col)
            
            # if best_score >= 0.85:
            #     mapped[excel_col] = best_col
            #     used_db.add(best_col)

            # elif 0.70 <= best_score < 0.85:
            #     ambiguous.append((excel_col, best_col, round(best_score, 3)))

            # else:
            #     unmapped.append((excel_col, round(best_score, 3)))

        return {
            "mapped": mapped,
            "ambiguous": ambiguous,
            "unmapped": unmapped
        }

    # =========================
    # ✅ Build config
    # =========================
    def build_config(self, file_path, key_columns):
        excel_info = self.get_excel_headers_with_color(file_path)

        headers_for_mapping = list(
            set(excel_info["update_headers"] + key_columns)
        )

        result = self.map(headers_for_mapping, key_columns)

        if result["ambiguous"]:
            raise Exception(f"Ambiguous mapping: {result['ambiguous']}")

        config = {
            "key_columns": key_columns,
            "header_mapping": result["mapped"]
        }

        logging.info(f"Mapping success: {len(result['mapped'])}")
        logging.info(f"Unmapped: {result['unmapped']}")

        return config