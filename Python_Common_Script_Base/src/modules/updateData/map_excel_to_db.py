from __future__ import annotations
import re
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)


def _normalize_text(text: Any) -> str:
    s = str(text or "").lower()
    return re.sub(r'[^a-z0-9]', '', s)


def _preprocess_for_embedding(text: Any) -> str:
    s = str(text or "").lower()
    s = re.sub(r'[_\-\.]', ' ', s)
    s = re.sub(r'[^a-z0-9\s]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


class ExcelHeaderExtractor:
    """Extracts headers from an Excel file and detects colored headers for update."""

    ALLOWED_EXT = {'.xlsx', '.xlsm', '.xltx', '.xltm'}

    def get_excel_headers_with_color(self, file_path: Path) -> Dict[str, List[str]]:
        if isinstance(file_path, (list, tuple)):
            file_path = file_path[0]

        file_path = Path(str(file_path))
        if file_path.suffix.lower() not in self.ALLOWED_EXT:
            raise ValueError(f"Unsupported Excel format: {file_path.suffix}")

        wb = load_workbook(file_path)
        ws = wb.active

        headers: List[str] = []
        colored: List[str] = []

        for col in ws.iter_cols(min_row=1, max_row=1):
            cell = col[0]
            h = str(cell.value).strip() if cell.value is not None else ""
            headers.append(h)

            fill = getattr(cell, 'fill', None)
            start = getattr(fill, 'start_color', None)
            if start is not None:
                idx = getattr(start, 'index', None)
                if idx and idx not in ("00000000", "FFFFFFFF"):
                    colored.append(h)

        wb.close()
        return {"all_headers": headers, "update_headers": colored}




# =============================================================================
# Pure-algorithm scorer — không dùng AI/ML/embedding
# =============================================================================

def _normalize(text: str) -> str:
    """Strip separators + special chars, lowercase."""
    text = str(text or "").lower()
    text = re.sub(r"[_\-\.\s/]", "", text)
    text = re.sub(r"[^a-z0-9]", "", text)
    return text


def _lcs_ratio(a: str, b: str) -> float:
    """
    Longest Common Subsequence / max(len(a), len(b)).
    Bắt thứ tự ký tự: "oppmk" vs "oppmoke" → 5/8 = 0.625
    O(m*n) — header thường < 40 ký tự nên không đáng kể.
    """
    if not a or not b:
        return 0.0
    m, n = len(a), len(b)
    # rolling two-row DP
    prev = [0] * (n + 1)
    curr = [0] * (n + 1)
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if a[i - 1] == b[j - 1]:
                curr[j] = prev[j - 1] + 1
            else:
                curr[j] = max(curr[j - 1], prev[j])
        prev, curr = curr, [0] * (n + 1)
    return prev[n] / max(m, n)


def _bigram_jaccard(a: str, b: str) -> float:
    """
    Jaccard trên tập bigram (cặp ký tự liên tiếp).
    Nhạy hơn char-set Jaccard: giữ ngữ cảnh vị trí.
    "opcode" bigrams: {op,pc,co,od,de}
    "operat" bigrams: {op,pe,er,ra,at} → giao {op} = 1/9 ≈ 0.11
    """
    if len(a) < 2 or len(b) < 2:
        # fallback về char Jaccard nếu quá ngắn
        sa, sb = set(a), set(b)
        union = sa | sb
        return len(sa & sb) / len(union) if union else 0.0
    bg_a = {a[i:i+2] for i in range(len(a) - 1)}
    bg_b = {b[i:i+2] for i in range(len(b) - 1)}
    union = bg_a | bg_b
    return len(bg_a & bg_b) / len(union) if union else 0.0


def _abbreviation_score(short: str, long: str) -> float:
    """
    Kiểm tra short có phải viết tắt của long không,
    bằng cách duyệt ký tự short theo thứ tự trong long.

    Trả về tỷ lệ ký tự của short khớp được / len(short).
    Có thưởng nếu khớp từ đầu (prefix weight).

    Ví dụ:
      short="op",   long="opcode"   → 2/2 = 1.0
      short="op",   long="operation"→ 2/2 = 1.0  (cả hai đều cao → dùng LCS phân biệt)
      short="pmk",  long="op_pmoke" → normalize → "oppmoke", khớp p-m-k → 3/3 = 1.0
      short="oppmk",long="oppmoke"  → khớp 5/5 = 1.0
    """
    if not short or not long:
        return 0.0
    # chỉ chạy nếu short ngắn hơn
    if len(short) >= len(long):
        return 0.0

    matched = 0
    pos = 0
    for ch in short:
        idx = long.find(ch, pos)
        if idx == -1:
            break
        matched += 1
        pos = idx + 1

    base = matched / len(short)

    # bonus: nếu long bắt đầu bằng short → viết tắt rõ ràng hơn
    prefix_len = 0
    for ca, cb in zip(short, long):
        if ca == cb:
            prefix_len += 1
        else:
            break
    prefix_bonus = min(prefix_len / len(short), 1.0) * 0.2

    return min(base + prefix_bonus, 1.0)


def _pair_score(excel_col: str, db_col: str) -> float:
    """Điểm tổng hợp cho 1 cặp (excel, db)."""
    a = _normalize(excel_col)
    b = _normalize(db_col)

    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    lcs     = _lcs_ratio(a, b)
    bigram  = _bigram_jaccard(a, b)

    # abbreviation: thử cả 2 chiều, lấy max
    abbrev  = max(
        _abbreviation_score(a, b),
        _abbreviation_score(b, a),
    )

    score = lcs * 0.45 + bigram * 0.30 + abbrev * 0.25
    return min(score, 1.0)


# =============================================================================
# HeaderMappingBuilder — drop-in replacement, xóa SentenceTransformer
# =============================================================================

class HeaderMappingBuilder:
    """
    Mapping Excel headers → DB columns bằng thuật toán thuần túy.
    Không dùng AI / ML / embedding.

    Mỗi excel_header lấy argmax độc lập trên toàn DB —
    đúng vì đề bài đảm bảo mỗi excel header có 1 DB column tương ứng
    và DB luôn có nhiều column hơn Excel.
    """

    def __init__(self, db_columns: List[str]):
        self.db_columns: List[str] = list(db_columns)
        # cache normalize để không tính lại nhiều lần
        self._db_norm: Dict[str, str] = {
            c: _normalize(c) for c in self.db_columns
        }

    # -----------------------------------------------------------------
    # Exact match (sau normalize)
    # -----------------------------------------------------------------
    def find_exact(
        self,
        excel_col: str,
        candidates: List[str],
    ) -> Optional[str]:
        norm = _normalize(excel_col)
        for c in candidates:
            if self._db_norm.get(c) == norm:
                return c
        return None

    # -----------------------------------------------------------------
    # Best match cho 1 excel header — argmax trên candidates
    # -----------------------------------------------------------------
    def find_best(
        self,
        excel_col: str,
        candidates: List[str],
    ) -> tuple[str, float]:
        """Trả về (db_col_tốt_nhất, score)."""
        best_col   = candidates[0] if candidates else ""
        best_score = -1.0

        for db_col in candidates:
            s = _pair_score(excel_col, db_col)
            if s > best_score:
                best_score = s
                best_col   = db_col

        return best_col, best_score

    # -----------------------------------------------------------------
    # Build config
    # -----------------------------------------------------------------
    def build_config(
        self,
        excel_headers: List[str],
        key_columns:   List[str],
    ) -> Dict[str, Any]:

        key_columns   = list(key_columns or [])
        header_mapping: Dict[str, str]   = {}
        scores:         Dict[str, float] = {}
        confidence:     Dict[str, str]   = {}
        needs_review:   List[str]        = []
        used_db:        set[str]         = set()

        # ----------------------------------------------------------
        # 1. Exact match cho key columns (ưu tiên tuyệt đối)
        # ----------------------------------------------------------
        for k in key_columns:
            match = self.find_exact(k, self.db_columns)
            if match:
                header_mapping[k] = match
                scores[k]         = 1.0
                confidence[k]     = "HIGH"
                used_db.add(match)
            else:
                logging.warning(f"Key column '{k}' not found in DB columns")

        # ----------------------------------------------------------
        # 2. Mỗi excel header còn lại → argmax độc lập
        #    (không dùng Hungarian vì DB >> Excel)
        # ----------------------------------------------------------
        remaining_excel = [h for h in excel_headers if h not in header_mapping]
        remaining_db    = [d for d in self.db_columns if d not in used_db]

        for excel_col in remaining_excel:

            # thử exact trước
            exact = self.find_exact(excel_col, remaining_db)
            if exact:
                header_mapping[excel_col] = exact
                scores[excel_col]         = 1.0
                confidence[excel_col]     = "HIGH"
                # KHÔNG add vào used_db — nhiều excel có thể map cùng 1 DB col
                continue

            best_col, best_score = self.find_best(excel_col, remaining_db)

            header_mapping[excel_col] = best_col
            scores[excel_col]         = round(best_score, 4)

            if best_score >= 0.75:
                confidence[excel_col] = "HIGH"
            elif best_score >= 0.55:
                confidence[excel_col] = "MED"
                needs_review.append(excel_col)
            else:
                confidence[excel_col] = "LOW"
                needs_review.append(excel_col)

        # ----------------------------------------------------------
        # 3. Fallback — header chưa được map (không nên xảy ra)
        # ----------------------------------------------------------
        for h in excel_headers:
            if h not in header_mapping:
                header_mapping[h] = ""
                scores[h]         = 0.0
                confidence[h]     = "LOW"
                needs_review.append(h)

        return {
            "key_columns":    key_columns,
            "header_mapping": header_mapping,
            "scores":         scores,
            "confidence":     confidence,
            "needs_review":   needs_review,
        }


class ExcelDBMapper:
    """Facade that combines header extraction and mapping builder.

    Public methods:
    - get_excel_headers_with_color(file_path)
    - map(file_path, key_columns)
    """

    def __init__(self, db_columns: List[str]):
        self.db_columns = list(db_columns)
        self.extractor = ExcelHeaderExtractor()
        self.builder = HeaderMappingBuilder(self.db_columns)

    def get_excel_headers_with_color(self, file_path: Path) -> Dict[str, List[str]]:
        return self.extractor.get_excel_headers_with_color(file_path)

    def map(self, file_path: Path, key_columns: Optional[List[str]] = None) -> Dict[str, Any]:
        key_columns = key_columns or []
        excel_info = self.extractor.get_excel_headers_with_color(file_path)
        excel_headers = list(set(excel_info.get('update_headers', []) + key_columns))
        return self.builder.build_config(excel_headers, key_columns)
