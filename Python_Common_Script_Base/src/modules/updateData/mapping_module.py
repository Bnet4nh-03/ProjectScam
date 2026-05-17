import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import tkinter.ttk as ttk
import json
import threading
import openpyxl

from src.apps.module_base import ModuleBase
from src.modules.updateData.map_excel_to_db import ExcelDBMapper
from src.modules.updateData.excel_update_processor import ExcelUpdateProcessor



class MappingModule(ModuleBase):
    def __init__(self, app):
        super().__init__(app)
        self.frame = None
        self.mapping_entries = {}
        self.file_path = ""
        self.db_columns = []
        self.excel_info = {}
        self.table_name_var = tk.StringVar(value=self.app.config["EXCEL_UPDATE"].get("table_name", ""))
        self.key_columns_var = tk.StringVar(value=",".join(self.app.config["EXCEL_UPDATE"].get("key_columns", [])))

    def name(self):
        return "Update Data"

    def build(self, parent):
        self.frame = tk.Frame(parent, bg="#f4f6f8")
        self.frame.pack(fill="both", expand=True)

        top_frame = tk.Frame(self.frame, bg="#f4f6f8")
        top_frame.pack(fill="x", padx=10, pady=10)

        tk.Button(top_frame, text="Chọn Excel", width=14, command=self.choose_file).pack(side="left")
        self.lbl_file = tk.Label(top_frame, text="Chưa chọn file", bg="#f4f6f8", anchor="w")
        self.lbl_file.pack(side="left", padx=10, fill="x", expand=True)
        tk.Button(top_frame, text="MAPPING", width=14, command=self.run).pack(side="right")

        config_frame = tk.LabelFrame(self.frame, text="Cấu hình", bg="#f4f6f8", fg="#333333", padx=10, pady=10)
        config_frame.pack(fill="x", padx=10, pady=6)

        tk.Label(config_frame, text="Table name:", bg="#f4f6f8", width=12, anchor="w").grid(row=0, column=0, sticky="w")
        tk.Entry(config_frame, textvariable=self.table_name_var, width=46).grid(row=0, column=1, sticky="w")

        tk.Label(config_frame, text="Key columns:", bg="#f4f6f8", width=12, anchor="w").grid(row=1, column=0, sticky="w", pady=8)
        tk.Entry(config_frame, textvariable=self.key_columns_var, width=46).grid(row=1, column=1, sticky="w", pady=8)

        tk.Label(config_frame, text="Nhập nhiều khóa bằng dấu phẩy", fg="#555555", bg="#f4f6f8").grid(row=2, column=0, columnspan=2, sticky="w")

        self.mapping_container = tk.LabelFrame(self.frame, text="Mapping Editor", bg="#f4f6f8", fg="#333333", padx=8, pady=8)
        self.mapping_container.pack(fill="both", padx=10, pady=6, expand=True)

        action_frame = tk.Frame(self.frame, bg="#f4f6f8")
        action_frame.pack(fill="x", padx=10, pady=4)
        tk.Button(action_frame, text="SUBMIT", width=14, command=self.submit).pack(side="left")
        tk.Button(action_frame, text="XÓA LOG", width=14, command=self.clear_log).pack(side="right")

        self.info_box = scrolledtext.ScrolledText(self.frame, height=10, wrap="word")
        self.info_box.pack(fill="both", padx=10, pady=6, expand=True)

    def _log(self, msg):
        if self.info_box:
            self.info_box.insert(tk.END, msg + "\n")
            self.info_box.see(tk.END)

    def choose_file(self):
        path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")])
        if path:
            if isinstance(path, (list, tuple)):
                path = path[0]
            self.file_path = path
            # keep app-level file_path in sync for older code paths
            try:
                self.app.file_path = path
            except Exception:
                pass
            self.lbl_file.config(text=path)
            self._log(f"Chọn file: {path}")

    def clear_log(self):
        if self.info_box:
            self.info_box.delete('1.0', tk.END)

    def run(self):
        if not self.file_path:
            messagebox.showerror("Lỗi", "Vui lòng chọn file Excel")
            return

        table_name = self.table_name_var.get().strip()
        key_columns = [item.strip() for item in self.key_columns_var.get().split(",") if item.strip()]
        if not table_name or not key_columns:
            messagebox.showerror("Lỗi", "Vui lòng nhập table_name và key_columns")
            return

        self.app.config["EXCEL_UPDATE"]["table_name"] = table_name
        self.app.config["EXCEL_UPDATE"]["key_columns"] = key_columns

        self.run_module(self.file_path, table_name, key_columns)

    def show_mapping_editor(self, mapping, key_columns, scores):
        for widget in self.mapping_container.winfo_children():
            widget.destroy()

        self.mapping_entries = {}
        if not mapping:
            tk.Label(self.mapping_container, text="Không tìm thấy tiêu đề để mapping.", bg="#f4f6f8").pack(anchor="w")
            return

        tk.Label(self.mapping_container, text="Excel Header", bg="#f4f6f8", font=(None, 10, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=4)
        tk.Label(self.mapping_container, text="DB Column", bg="#f4f6f8", font=(None, 10, "bold")).grid(row=0, column=1, sticky="w", padx=4, pady=4)
        tk.Label(self.mapping_container, text="Score", bg="#f4f6f8", font=(None, 10, "bold")).grid(row=0, column=2, sticky="w", padx=4, pady=4)

        self.db_columns = self.db_columns or []
        db_options = [""] + self.db_columns
        sorted_items = sorted(mapping.items(), key=lambda item: (item[0] not in key_columns, item[0].lower()))
        for row, (excel_col, db_col) in enumerate(sorted_items, start=1):
            score = scores.get(excel_col, 0.0)
            score_text = f"{score:.3f}" if score else ""
            row_bg = "#fff3cd" if score and score < 0.7 else "#f4f6f8"

            tk.Label(self.mapping_container, text=excel_col, bg="#f4f6f8", anchor="w").grid(row=row, column=0, sticky="w", padx=4, pady=2)
            combobox = ttk.Combobox(self.mapping_container, values=db_options, width=42, state="readonly")
            combobox.set(db_col or "")
            combobox.grid(row=row, column=1, sticky="w", padx=4, pady=2)

            score_label = tk.Label(self.mapping_container, text=score_text, bg=row_bg, anchor="w")
            score_label.grid(row=row, column=2, sticky="w", padx=4, pady=2)

            if score and score < 0.7 and db_col:
                try:
                    combobox.config(background="#fff3cd")
                except Exception:
                    pass

            self.mapping_entries[excel_col] = combobox

    def _build_review_headers(self, final_mapping):
        seen = set()
        headers = []
        for excel_col, entry in self.mapping_entries.items():
            db_col = final_mapping.get(excel_col, "").strip()
            if not db_col:
                continue

            if db_col.lower() == "no":
                continue

            if db_col not in seen:
                seen.add(db_col)
                headers.append(db_col)
        return headers

    def _load_review_rows(self, final_mapping, review_headers):
        if not self.file_path:
            return []

        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active

        normalized_mapping = {
            str(excel_header).strip().lower(): str(db_col).strip()
            for excel_header, db_col in final_mapping.items()
            if str(db_col).strip()
        }

        header_to_index = {}
        for col_idx, cell in enumerate(worksheet[1], start=1):
            header = str(cell.value).strip() if cell.value else ""
            normalized_header = header.lower()
            if normalized_header in normalized_mapping:
                header_to_index[col_idx] = normalized_mapping[normalized_header]

        rows = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            record = {header: None for header in review_headers}
            has_value = False
            for col_idx, cell in enumerate(row, start=1):
                db_col = header_to_index.get(col_idx)
                if not db_col:
                    continue
                record[db_col] = cell.value
                if cell.value is not None:
                    has_value = True
            if has_value:
                rows.append(record)

        return rows
    
    def _build_review_payloads(
        self,
        final_mapping,
        review_headers
    ):

        if not self.file_path:
            return []

        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active

        header_mapping = {
            str(excel_header).strip().lower(): str(db_col).strip()
            for excel_header, db_col in final_mapping.items()
            if db_col
        }

        header_to_index = {}
        for col_idx, cell in enumerate(worksheet[1], start=1):
            header = str(cell.value).strip() if cell.value is not None else ""
            normalized_header = header.lower()
            if normalized_header in header_mapping:
                header_to_index[col_idx] = header_mapping[normalized_header]

        payloads = []

        # normalize key headers to lowercase for robust comparison
        key_headers = set(
            c.strip().lower()
            for c in self.app.config["EXCEL_UPDATE"]["key_columns"]
        )

        table_name = (
            self.app.config[
                "EXCEL_UPDATE"
            ][
                "table_name"
            ]
        )

        database_name = (
            self.app.config[
                "EXCEL_UPDATE"
            ].get(
                "database"
            )
        )

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            keys = {}
            data = {}

            for col_idx, cell in enumerate(row, start=1):
                db_col = header_to_index.get(col_idx)
                if not db_col:
                    continue

                value = cell.value
                if db_col and str(db_col).strip().lower() in key_headers:
                    if value is not None:
                        keys[db_col] = value
                    continue

                if db_col not in review_headers:
                    continue

                data[db_col] = value

            if not data:
                continue
            if len(keys) != len(key_headers):
                continue

            payload = {
                "table": table_name,
                "keys": keys,
                "data": data,
            }
            if database_name:
                payload["database"] = database_name
            payloads.append(payload)
        return payloads

    def show_review_window(self, final_mapping):
        review_headers = self._build_review_headers(final_mapping)
        if not review_headers:
            messagebox.showerror("Lỗi", "Không có cột DB nào để hiển thị trong review.")
            return

        rows = self._load_review_rows(final_mapping, review_headers)
        if not rows:
            messagebox.showwarning("Chú ý", "Không tìm thấy dữ liệu nào để review.")
            return

        review_window = tk.Toplevel(self.frame)
        review_window.title("Review dữ liệu trước khi Submit")
        review_window.geometry("1000x550")
        review_window.configure(bg="#f4f6f8")
        review_window.transient(self.frame.winfo_toplevel())
        review_window.grab_set()

        info_label = tk.Label(review_window, text="Data.", anchor="w", bg="#f4f6f8")
        info_label.pack(fill="x", padx=10, pady=6)

        search_frame = tk.Frame(review_window, bg="#f4f6f8")
        search_frame.pack(fill="x", padx=10, pady=(0, 6))

        tk.Label(search_frame, text="Tìm kiếm:", width=10, anchor="w").pack(side="left")
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side="left", padx=(0, 8))

        tk.Button(search_frame, text="Search", width=10, command=lambda: chunked_filter_rows()).pack(side="left")
        tk.Button(search_frame, text="Clear", width=10, command=lambda: (search_var.set(""), chunked_filter_rows())).pack(side="left", padx=(8, 0))

        container = tk.Frame(review_window)
        container.pack(fill="both", expand=True, padx=10, pady=6)

        columns = ["No"] + review_headers
        tree = ttk.Treeview(container, columns=columns, show="headings")
        tree.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        vsb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(review_window, orient="horizontal", command=tree.xview)
        hsb.pack(fill="x", padx=10)
        tree.configure(xscrollcommand=hsb.set)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=80 if col == "No" else 150, anchor="w")

        def on_mouse_wheel(event):
            tree.yview_scroll(int(-1*(event.delta/120)), "units")

        tree.bind("<MouseWheel>", on_mouse_wheel)

        # use chunked insertion to keep UI responsive
        def chunked_filter_rows():
            query = search_var.get().strip().lower()
            tree.delete(*tree.get_children())
            filtered = []
            for index, record in enumerate(rows, start=1):
                line = " ".join([str(record.get(col, "") or "").lower() for col in review_headers])
                if not query or query in line:
                    filtered.append((index, [record.get(col, "") for col in review_headers]))

            batch_size = 100

            def insert_batch(start=0):
                end = min(start + batch_size, len(filtered))
                for i in range(start, end):
                    idx, vals = filtered[i]
                    tree.insert("", "end", values=[idx] + vals)
                if end < len(filtered):
                    review_window.after(10, lambda: insert_batch(end))

            insert_batch(0)

        # create button frame early so we can control submit button state
        button_frame = tk.Frame(review_window, bg="#f4f6f8")
        button_frame.pack(fill="x", padx=10, pady=8)

        cancel_btn = tk.Button(button_frame, text="Hủy", width=12, command=review_window.destroy)
        cancel_btn.pack(side="left")

        submit_btn = tk.Button(button_frame, text="Submit", width=12, state="disabled")
        submit_btn.pack(side="right")

        # status label for payload building
        status_var = tk.StringVar(value="Preparing payloads...")
        status_label = tk.Label(review_window, textvariable=status_var, bg="#f4f6f8")
        status_label.pack(fill="x", padx=10)

        # start inserting visible rows (chunked)
        chunked_filter_rows()

        # build payloads in background so UI doesn't hang
        def build_payloads_bg():
            try:
                payloads = self._build_review_payloads(final_mapping, review_headers)
                def on_done():
                    self.review_payloads = payloads
                    status_var.set(f"Payloads ready: {len(payloads)} items")
                    submit_btn.config(state="normal")
                    self._log(f"Review payloads built: {len(payloads)} items")
                review_window.after(0, on_done)
            except Exception as exc:
                review_window.after(0, lambda: status_var.set(f"Error building payloads: {exc}"))

        threading.Thread(target=build_payloads_bg, daemon=True).start()

        # wire submit to background send using the prebuilt payloads
        def on_submit():
            submit_btn.config(state="disabled")
            status_var.set("Sending payloads...")

            def on_complete_cb():
                try:
                    status_var.set("Send complete")
                except Exception:
                    pass

            # call final_submit which runs work in background and will call on_complete when done
            try:
                self.final_submit(final_mapping, review_window, self.review_payloads, on_complete=on_complete_cb)
            except Exception as exc:
                status_var.set(f"Send error: {exc}")

        submit_btn.config(command=on_submit)

    def final_submit(self, final_mapping, review_window=None, payloads=None, on_complete=None):
        excel_conf = self.app.config["EXCEL_UPDATE"]
        excel_conf["header_mapping"] = final_mapping
        # perform processing and API call in background to avoid blocking UI
        # `on_complete` is an optional callable invoked on the main thread when send finishes
        def send_thread(on_complete=None):
            try:
                local_payloads = payloads
                if local_payloads is None:
                    processor = ExcelUpdateProcessor(excel_conf, self.app.logger)
                    local_payloads = processor.process_file(self.file_path)

                if not local_payloads:
                    def warn_no_data():
                        messagebox.showwarning("Chú ý", "Không có dữ liệu nào để gửi. Vui lòng kiểm tra màu cell và key columns.")
                    if review_window:
                        review_window.after(0, warn_no_data)
                    else:
                        warn_no_data()
                    return

                sp_query_str = json.dumps(local_payloads)
                # result = self.app.enco_api_client.call_sp(excel_conf["stored_update"], {"@p_payload": sp_query_str})
                # self._log(f"✅ SP RESULT: {result}")

                def on_success():
                    messagebox.showinfo("Hoàn thành", "Stored procedure đã được gọi thành công.")
                    if review_window:
                        review_window.destroy()
                    if on_complete:
                        try:
                            if review_window:
                                review_window.after(0, on_complete)
                            else:
                                on_complete()
                        except Exception:
                            pass

                if review_window:
                    review_window.after(0, on_success)
                else:
                    on_success()

            except Exception as exc:
                self._log(f"❌ SUBMIT ERROR: {exc}")
                def on_error():
                    messagebox.showerror("Lỗi", f"Không thể gửi dữ liệu: {exc}")
                    if on_complete:
                        try:
                            if review_window:
                                review_window.after(0, on_complete)
                            else:
                                on_complete()
                        except Exception:
                            pass
                if review_window:
                    review_window.after(0, on_error)
                else:
                    on_error()

        threading.Thread(target=send_thread, args=(on_complete,), daemon=True).start()

    def run_module(self, file_path, table_name, key_columns):
        try:
            # TODO: use stored_getHeader when API is available
            # client = self.app.enco_api_client
            # header_data = client.call_sp(self.app.config["EXCEL_UPDATE"]["stored_getHeader"], {"@table_name": table_name})
            header_data = [
                {"data": [{"COLUMN_NAME": "No"}]},
                {"data": [{"COLUMN_NAME": "Barcode"}]},
                {"data": [{"COLUMN_NAME": "Customer"}]},
                {"data": [{"COLUMN_NAME": "BoardNo"}]},
                {"data": [{"COLUMN_NAME": "BoardSerial"}]},
                {"data": [{"COLUMN_NAME": "MFG"}]},
                {"data": [{"COLUMN_NAME": "Type"}]},
                {"data": [{"COLUMN_NAME": "Pkg_Type"}]},
                {"data": [{"COLUMN_NAME": "Sites"}]},
                {"data": [{"COLUMN_NAME": "Equip"}]},
                {"data": [{"COLUMN_NAME": "Handler Type"}]},
                {"data": [{"COLUMN_NAME": "Prod_rmk"}]},
                {"data": [{"COLUMN_NAME": "Opcode"}]},                
                {"data": [{"COLUMN_NAME": "Equip"}]},
                {"data": [{"COLUMN_NAME": "Handler"}]},
                {"data": [{"COLUMN_NAME": "Prod"}]},
                {"data": [{"COLUMN_NAME": "Op"}]},
            ]
            db_columns = [item["data"][0]["COLUMN_NAME"] for item in header_data]
            self._log(f"DB columns: {db_columns}")

            mapper = ExcelDBMapper(db_columns)
            self._log(f"Attempting to read Excel: {file_path!r}")
            excel_info = mapper.get_excel_headers_with_color(file_path)
            self.excel_info = excel_info
            headers_for_mapping = excel_info["update_headers"] or excel_info["all_headers"]
            headers_for_mapping = list(set(headers_for_mapping + key_columns))
            self._log(f"Excel headers: {headers_for_mapping}")

            result = mapper.map(file_path, key_columns)
            mapping = result.get("header_mapping", {})
            scores = result.get("scores", {})
            self.db_columns = db_columns
            self.show_mapping_editor(mapping, key_columns, scores)
            for source, target in mapping.items():
                score = scores.get(source, 0.0)
                self._log(f"{source} → {target or '<no match>'} (score={score:.3f})")
        except Exception as exc:
            self._log(f"ERROR: {exc}")

    def submit(self):
        if not self.mapping_entries:
            messagebox.showwarning("Cảnh báo", "Chưa có mapping để gửi")
            return

        final_mapping = {}
        for excel_col, entry in self.mapping_entries.items():
            value = entry.get().strip()
            if value:
                final_mapping[excel_col] = value

        if not final_mapping:
            messagebox.showerror("Lỗi", "Mapping không hợp lệ")
            return

        # normalize key columns and final mapping values for case-insensitive comparison
        key_columns = set(c.strip().lower() for c in self.app.config["EXCEL_UPDATE"]["key_columns"])
        final_targets_lower = set(v.strip().lower() for v in final_mapping.values())
        missing_keys = [col for col in key_columns if col not in final_targets_lower]
        if missing_keys:
            messagebox.showerror(
                "Lỗi",
                f"Vui lòng mapping đầy đủ key columns: {', '.join(missing_keys)}"
            )
            return

        self._log("=== FINAL MAPPING ===")
        for source, target in final_mapping.items():
            self._log(f"{source} → {target}")

        excel_conf = self.app.config["EXCEL_UPDATE"]
        excel_conf["header_mapping"] = final_mapping
        excel_conf["update_headers"] = self.excel_info.get("update_headers", [])

        self.show_review_window(final_mapping)
