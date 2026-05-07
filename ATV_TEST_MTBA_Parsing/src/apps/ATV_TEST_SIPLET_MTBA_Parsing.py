from src.core.base_app import BaseApp
from src.utility.enco_common_api_utils.enco_api_client import EnCoApiClient
from datetime import datetime, timezone, timedelta
import os
import csv
import requests
import shutil
import sqlite3
from concurrent.futures import ThreadPoolExecutor
import json
import re
import threading
from openpyxl import load_workbook
import tempfile

sessions = [requests.Session() for _ in range(10)]

class SipletMTBAParsing(BaseApp):
    def __init__(self, config_path: str, config_type: str):
        super().__init__(self.__class__.__name__, config_path, config_type)
        self.enco_api_client = EnCoApiClient(self.config["API_Common"].get("api_link_2"), self.logger)
        self.local_dir = self.config["Path"].get("Local")
        self.alarm_lock = threading.Lock()   
        self.error_dict = {}

    #Add the error list to the db
    def add_event_error(self, path_file = r"\JAM_list\Siplet\Siplet_Jam_list.xlsx"):
        wb = load_workbook(filename=path_file, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = set()
        
            for r in range(2, ws.max_row + 1):  # bỏ header
                jam_code = ws.cell(row=r, column=1).value
                message  = ws.cell(row=r, column=2).value

                if not jam_code or not message:
                    continue

                message = str(message).strip()
                jam_code = str(jam_code).strip()

                try:
                    #create query to add error
                    json_obj = {
                        "rcs_type": sheet_name,
                        "event_type": "JAM",
                        "event_code": jam_code,
                        "event_message": message,
                        "created_by": "250707",
                        "active_flag": 1
                    }
                    sp_query_str = json.dumps(json_obj, ensure_ascii=False)
                    self.enco_api_client.call_sp('TestDB..USP_PmMonitor_SetEventInfo', {"@json_para": sp_query_str}) 
                    self.logger.info(f"Machine {sheet_name}: Add OK: event_code: {jam_code}, message: {message}") 
                    result.append(message)
                except Exception as e:
                    self.logger.info(f"Machine {sheet_name}: Add FAILED: event_code: {jam_code}, message: {message}")
                    continue
            else:
                self.logger.info(f"Machine {sheet_name}: [{jam_code} _ {message}] existed")

    #Get list error IDs by machine type
    def get_error_dict(self):
        file_path = "Siplet_error_dict.txt"
        if not os.path.exists(file_path):
            list_rcs = ['HT-3016S', 'HT-3016S+', 'TWSL421', 'TW154V']
            lines = []

            for rcs in list_rcs:
                sp_query = {"rcs_type": rcs, "event_type": "JAM", "active_flag": 1}
                sp_query_str = json.dumps(sp_query, ensure_ascii=False)  # Convert dict to JSON string

                #get list error in db
                result = self.enco_api_client.call_sp('TestDB..USP_PmMonitor_GetEventInfo', {"@json_para": sp_query_str})
                info = result[0]
                info_data = info['data']
                event_dict = {item['event_id'] : item['event_message'] for item in info_data}
                
                for event_id, message in sorted(event_dict.items()):
                    lines.append(f"{rcs} : {event_id} : {message}")

            #write list error IDs into error dict  
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")

    # create column last_marker to save reading milestone in Qorvo_Machine_Define.db
    def create_last_marker(self, db_path):           
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("ALTER TABLE Siplet_Machine_Define ADD COLUMN last_marker TEXT")
        except sqlite3.OperationalError:
            pass
        
        cursor.execute("""
            UPDATE Siplet_Machine_Define 
               SET last_marker = ?
        """, ('',))

        conn.commit()
        conn.close()

    #save reading milestone to db
    def set_last_marker(self, db_path, machine_name, last_marker):
        try:
            conn = sqlite3.connect(db_path, timeout=30)
            cursor = conn.cursor()      
            cursor.execute("""
                UPDATE Siplet_Machine_Define 
                SET last_marker = ?
                WHERE machine_name = ?
            """, (last_marker, machine_name))

            conn.commit()
            conn.close() 
        except Exception as e:
            self.logger.error(f"Fail update last marker {machine_name}: {e}")   

    #read sub_file for HT-9045F machine
    def get_log_file(self, input_dir, pattern, machine_type):
        if machine_type == "HT3016":
            for entry in sorted(os.listdir(input_dir)):
                if entry == "Handler.db3":
                    file_path = os.path.join(input_dir, entry)
                    # self.logger.info(file_path)
                    if os.path.isfile(file_path):
                        return file_path

        else:
            subfolder_name = datetime.now().strftime("%Y%m%d")
            subfolder_path = os.path.join(input_dir, subfolder_name)
            for entry in sorted(os.listdir(subfolder_path)):
                file_path = os.path.join(subfolder_path, entry)
                # self.logger.info(file_path)
                if os.path.isfile(file_path):
                    return file_path
        return None

    
    def load_error_dict(self, file_path="Siplet_error_dict.txt"):
        """Load file error_dict.txt vào dictionary trong RAM."""

        with self.alarm_lock:
            self.error_dict = {}  # reset

            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()  # loại bỏ \n, space

                        if not line:
                            continue

                        part = line.split(" : ")
                        if len(part) != 3:
                            # self.logger.warning(f"Format sai: {line}")
                            continue

                        machine_type = part[0].strip()
                        event_id = part[1].strip()
                        message = part[2].strip().lower()

                        # tạo key machine nếu chưa có
                        if machine_type not in self.error_dict:
                            self.error_dict[machine_type] = {}

                        # ánh xạ message → ID
                        self.error_dict[machine_type][message] = event_id

            except Exception as e:
                self.logger.error(f"Cannot load error dict: {e}")


    #get error id by message in error dict
    def get_error_id(self, message, machine_type):
        message_str = str(message).strip().lower()

        with self.alarm_lock:  
            if machine_type not in self.error_dict:
                return None
            
            return self.error_dict[machine_type].get(message_str, None)
        
    #read log to get error in testing
    def get_error_stack(self, file, db_path, machine_name, mnbr, machine_type, last_marker):
        stack = [] 

        if last_marker:
            last_marker_date = datetime.strptime(last_marker, "%Y-%m-%d %H:%M:%S")
        else:
            last_marker_date = None

        #read logs by device type
        if machine_type == "HT-3016" or machine_type == "'HT-3016S+":
            try:
                # print(file)
                valid_entries = []

                conn = sqlite3.connect(file, timeout=30)
                conn.text_factory = bytes
                cursor = conn.cursor()

                query = "SELECT OccurDateTime, Message FROM EventLog"

                if last_marker:
                    query += f" WHERE OccurDateTime > '{last_marker}'"
                    
                cursor.execute(query)
                entries = cursor.fetchall()
                conn.close()

                for time_str, message in entries:
                    if not isinstance(message, bytes):
                        continue
                    try:
                        time_str = time_str.decode("utf-8")
                        message = message.decode("utf-8")     
                    except:
                        continue

                    valid_entries.append((time_str, message))

                for time_str, message in valid_entries:    
                    id_error = self.get_error_id(message, machine_type) #get error id from error dict

                    if not id_error:
                        continue
                    
                    stack.append((machine_name, mnbr, time_str, id_error)) #add error to stack
                    self.set_last_marker(
                        db_path,
                        machine_name,
                        time_str
                    ) #update last_marker to Qorvo_Machine_Define.db

            except Exception as e:
                self.logger.info(f"Machine {machine_name}: {e}")    

        if machine_type == 'TWSL421':
            try:
                with open(file, "r", encoding="latin1 ", newline="") as f:
                    for line in f:
                        message = None

                        parts = line.split("\t")
                        message = parts[1].split("ENGLISH =")[1].strip()
                        error_english = message.split(", CHINA")[0].strip()

                        id_error = self.get_error_id(error_english, machine_type)         #get error id from error dict
                        
                        if not id_error:
                            continue

                        datetime_event_str = parts[0].split(".")[0].strip()
                        # print(f"{datetime_event_str} : {id_error} : {error_english}")
                        datetime_event = datetime.strptime(datetime_event_str, "%Y-%m-%d %H:%M:%S")

                        if last_marker_date is None or datetime_event > last_marker_date:
                            stack.append((machine_name, mnbr, datetime_event_str, id_error))       #add error to stack
                            self.set_last_marker(
                                db_path,
                                machine_name,
                                datetime_event.strftime("%Y/%m/%d %H:%M:%S")
                            ) #update last_marker to Qorvo_Machine_Define.db
                        else:
                            continue 
            except Exception as e:
                self.logger.info(f"Machine {machine_name}: {e}")   
 
        else:
            try:
               with open(file, "r", encoding="latin1", newline="") as f:
                    for line in f:
                        if not line[:2].isdigit() or line[2] != '/':
                            continue

                        if " : " not in line:
                            continue

                        try:
                            raw_date = line.split(" ",2)[0]
                            raw_time = line.split(" ", 2)[1]

                            date_str = f"{datetime.now().year}-{raw_date.replace('/', '-')}"
                            time_str = raw_time.split(".")[0]
                            timestamp = f"{date_str} {time_str}"
                            error_message = line.split(" : ", 1)[1].strip()
                        except:
                            continue
                        
                        id_error = self.get_error_id(error_message, machine_type) #get error id from error dict
                        if not id_error:
                            continue

                        datetime_event = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")

                        if last_marker_date is None or datetime_event > last_marker_date:
                            stack.append((machine_name, mnbr, timestamp, id_error))        #add error to stack                                                                                                                                                                                                                                                                        
                            self.set_last_marker(
                                db_path,
                                machine_name,
                                timestamp
                            ) #update last_marker to Qorvo_Machine_Define.db
            except Exception as e:
                self.logger.info(f"Machine {machine_name}: {e}")    
        return stack

    #add error to database
    def parsing_ULT(self, db_path, machine_name, input_dir, pattern, mnbr, machine_type, last_marker):
        file = self.get_log_file(input_dir, pattern, machine_type)        #Get the log files for today
        if not file:
            # self.logger.info(f"Machine {machine_name} can not get file")  
            return

        temp_dir_ctx = tempfile.TemporaryDirectory(prefix="ult_log_", ignore_cleanup_errors=True)
        temp_dir = temp_dir_ctx.name
        dst_file = os.path.join(temp_dir, os.path.basename(file))
        
        shutil.copy2(file, dst_file)
        stack = self.get_error_stack(dst_file, db_path, machine_name, mnbr, machine_type, last_marker)      #get error in log files

        if stack:
            #create query to add error to database
            data = self.create_query(stack)    
            sp_query = {
                "temp_table": "@",
                "data" : data
            }      
            sp_query_str = json.dumps(sp_query, ensure_ascii=False)
            # print(sp_query_str)
        #     #execute query
            self.enco_api_client.call_sp('TestDB..USP_PmMonitor_SetMachineLogs', {"@json_para": sp_query_str})

    #create query to insert data to db
    def create_query(self, stack): 
        result = []
        for machine_name, mnbr, datetime_event_str, id_error in stack:
            datetime_event_local, datetime_event_utc = self.convert_time_utc(datetime_event_str)
            try:
                event_time = datetime_event_local.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3]     #event time in local timezone
                event_time_utc = datetime_event_utc.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"     #event time in utc timezone
                result.append({
                    "mnbr": mnbr,       #code machine
                    "event_time": event_time,       #event time in local timezone
                    "event_time_utc": event_time_utc,       #event time in utc timezone
                    "event_id": id_error
                })
                self.logger.info(f"Machine {machine_name}: Add OK: mnbr: {mnbr}, event_time:{event_time}, event_time_utc: {event_time_utc}, event_id: {id_error}")
            except Exception as e:
                self.logger.info(f"Machine {machine_name}: Add FAILED: mnbr: {mnbr}, event_time:{event_time}, event_time_utc: {event_time_utc}, event_id: {id_error}")
                continue
        return result
    
    #convert time in local tz to utc tz
    def convert_time_utc(self, event_time_str):
        try:
            dt_event = datetime.strptime(event_time_str, "%Y-%m-%d %H:%M:%S")
            event_local = dt_event.replace(tzinfo=timezone(timedelta(hours=7)))
            event_utc = event_local.astimezone(timezone.utc)
            return event_local, event_utc
        except Exception as e:
            self.logger.info(f"{e}")   

    # def set_last_marker_2(self, db_path, machine_name, last_marker):
    #         try:
    #             conn = sqlite3.connect(db_path, timeout=30)
    #             cursor = conn.cursor()      
    #             cursor.execute("""
    #                 UPDATE Siplet_Machine_Define 
    #                 SET last_marker = ?
    #                 WHERE machine_name = ?
    #             """, (last_marker, machine_name))

    #             conn.commit()
    #             conn.close() 
    #         except Exception as e:
    #             self.logger.error(f"Fail update last marker {machine_name}: {e}")   

    def run(self):
        self.logger.info(f"Starting {self.__class__.__name__} application")
                                       
        # add new event error to db
        # self.add_event_error()
        self.get_error_dict()

        self.load_error_dict()

        # #get Qorvo_Machine_Define.db file containing local information
        db_path = os.path.join(self.local_dir, "Siplet_Machine_Define.db")
        copy_db_path = "Siplet_Machine_Define.db"
        if not os.path.exists(copy_db_path):
            shutil.copyfile(db_path, copy_db_path)
            self.create_last_marker(copy_db_path)       #create column last_marker

        # #get data in Qorvo_Machine_Define.db
        conn = sqlite3.connect(copy_db_path, timeout=30)
        cursor = conn.cursor()
        cursor.execute("SELECT machine_name, input_dir, pattern, mnbr, machine_type, last_marker FROM Siplet_Machine_Define")
        entries = cursor.fetchall()
        conn.close()

        # self.set_last_marker_2(copy_db_path, "TS-HT3#002", "2026-02-01 00:00:00")

        #run multithread
        def process_local_entry(entry):
                machine_name, input_dir, pattern, mnbr, machine_type, last_marker = entry
                last_marker = last_marker or ''
                self.parsing_ULT(copy_db_path, machine_name, input_dir, pattern, mnbr, machine_type, last_marker) 

        with ThreadPoolExecutor(max_workers=6) as executor:
            executor.map(process_local_entry, entries) 

        self.logger.info(f"{self.__class__.__name__} application finished")

