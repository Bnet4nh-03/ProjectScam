from src.core.base_app import BaseApp
from src.utility.enco_common_api_utils.enco_api_client import EnCoApiClient
from datetime import datetime, timezone, timedelta
import os
import csv
import requests
import shutil
import sqlite3
from concurrent.futures import ThreadPoolExecutor
import json
import re
import threading
from openpyxl import load_workbook
import tempfile

sessions = [requests.Session() for _ in range(10)]

class HexaMTBAParsing(BaseApp):
    def __init__(self, config_path: str, config_type: str): 
        super().__init__(self.__class__.__name__, config_path, config_type)
        self.enco_api_client = EnCoApiClient(self.config["API_Common"].get("api_link_2"), self.logger)
        self.local_dir = self.config["Path"].get("Local")
        self.alarm_lock = threading.Lock()   
        self.error_dict = {}

    #Add the error list to the db
    def add_event_error(self, path_file = r"JAM_list\Hexa\Hexa_Jam_list.xlsx"):
        wb = load_workbook(filename=path_file, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # result = set()
        
            for r in range(2, ws.max_row + 1):  # bỏ header
                jam_code = ws.cell(row=r, column=1).value
                message  = ws.cell(row=r, column=2).value

                if not jam_code or not message:
                    continue

                message = str(message).strip()
                jam_code = str(jam_code).strip()

                try:
                    #create query to add error
                    json_obj = {
                        "rcs_type": sheet_name,
                        "event_type": "JAM",
                        "event_code": jam_code,
                        "event_message": message,
                        "created_by": "250707",
                        "active_flag": 1
                    }
                    sp_query_str = json.dumps(json_obj, ensure_ascii=False)
                    self.enco_api_client.call_sp('TestDB..USP_PmMonitor_SetEventInfo', {"@json_para": sp_query_str}) 
                    self.logger.info(f"Machine {sheet_name}: Add OK: event_code: {jam_code}, message: {message}") 
                    # result.add(message)
                except Exception as e:
                    self.logger.info(f"Machine {sheet_name}: Add FAILED: event_code: {jam_code}, message: {message}")
                    continue
            else:
                self.logger.info(f"Machine {sheet_name}: [{jam_code} _ {message}] existed")

    #Get list error IDs by machine type
    def get_error_dict(self):
        file_path = "Hexa_error_dict.txt"
        if not os.path.exists(file_path):
            list_rcs = ['HEXA EVO+']
            lines = []

            for rcs in list_rcs:
                sp_query = {"rcs_type": rcs, "event_type": "JAM", "active_flag": 1}
                sp_query_str = json.dumps(sp_query, ensure_ascii=False)  # Convert dict to JSON string

                #get list error in db
                result = self.enco_api_client.call_sp('TestDB..USP_PmMonitor_GetEventInfo', {"@json_para": sp_query_str})
                info = result[0]
                info_data = info['data']
                event_dict = {item['event_id'] : item['event_message'] for item in info_data}
                
                for event_id, message in sorted(event_dict.items()):
                    lines.append(f"{rcs} : {event_id} : {message}")

            #write list error IDs into error dict  
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")

    # create column last_marker to save reading milestone in HEXA_machine_define.db
    def create_last_marker(self, db_path):           
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("ALTER TABLE HEXA_machine_define ADD COLUMN last_marker TEXT")
        except sqlite3.OperationalError:
            pass
        
        cursor.execute("""
            UPDATE HEXA_machine_define
               SET last_marker = ?
        """, ('',))

        conn.commit()
        conn.close()

    #save reading milestone to db
    def set_last_marker(self, db_path, machine_name, last_marker):
        try:
            conn = sqlite3.connect(db_path, timeout=30)
            cursor = conn.cursor()      
            cursor.execute("""
                UPDATE HEXA_machine_define
                SET last_marker = ?
                WHERE machine_name = ?
            """, (last_marker, machine_name))

            conn.commit()
            conn.close() 
        except Exception as e:
            self.logger.error(f"Fail update last marker {machine_name}: {e}")   

    #get log file by day
    def get_log_file(self, input_dir):
        file_name = datetime.now().strftime("%Y%m%d")

        for entry in sorted(os.listdir(input_dir)):
            if file_name in entry:
                file_path = os.path.join(input_dir, entry)
                # self.logger.info(file_path)
                if os.path.isfile(file_path):
                    return file_path
        return None

    
    def load_error_dict(self, file_path="Hexa_error_dict.txt"):
        """Load file error_dict.txt vào dictionary trong RAM."""

        with self.alarm_lock:
            self.error_dict = {}  # reset

            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()  # loại bỏ \n, space

                        if not line:
                            continue

                        part = line.split(" : ")
                        if len(part) != 3:
                            # self.logger.warning(f"Format sai: {line}")
                            continue

                        machine_type = part[0].strip()
                        event_id = part[1].strip()
                        message = part[2].strip().lower()

                        # tạo key machine nếu chưa có
                        if machine_type not in self.error_dict:
                            self.error_dict[machine_type] = {}

                        # ánh xạ message → ID
                        self.error_dict[machine_type][message] = event_id

            except Exception as e:
                self.logger.error(f"Cannot load error dict: {e}")


    #get error id by message in error dict
    def get_error_id(self, message, machine_type):
        message_str = str(message).strip().lower()

        with self.alarm_lock:  
            if machine_type not in self.error_dict:
                return None
            
            return self.error_dict[machine_type].get(message_str, None)
        
    #read log to get error in testing
    def get_error_stack(self, file, db_path, machine_name, mnbr, machine_type, last_marker):
        stack = [] 

        if last_marker:
            last_marker_date = datetime.strptime(last_marker, "%d/%m/%Y %H:%M:%S")
        else:
            last_marker_date = None

        #read logs by device type
        if machine_type == "HEXA EVO+":
            try:
                with open(file, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()

                        if not line or line.startswith(";"):
                            continue

                        parts = line.split(",")
                        message  = ",".join(parts[6:]) 
                        clean_message = re.sub(r"<.*?>", "", message).strip()

                        id_error = self.get_error_id(clean_message, machine_type)
                        # id_error = 123
                        if id_error:
                            date_raw = parts[0].strip() 
                            time_raw = parts[1].strip()

                            datetime_event_str = f"{date_raw} {time_raw}"

                            datetime_event = datetime.strptime(datetime_event_str, "%d/%m/%Y %H:%M:%S")

                            if last_marker_date is None or datetime_event > last_marker_date:
                                stack.append((machine_name, mnbr, datetime_event_str, id_error))
                                self.set_last_marker(
                                    db_path,
                                    machine_name,
                                    datetime_event.strftime("%d/%m/%Y %H:%M:%S")
                                )
                            else:
                                continue
            except Exception as e:
                self.logger.info(f"Machine {machine_name}: {e}")    

        return stack

    #add error to database
    def parsing_ULT(self, db_path, machine_name, input_dir, pattern, mnbr, machine_type, last_marker):
        file = self.get_log_file(input_dir)        #Get the log files for today
        if not file:
            return

        temp_dir_ctx = tempfile.TemporaryDirectory(prefix="ult_log_", ignore_cleanup_errors=True)
        temp_dir = temp_dir_ctx.name
        dst_file = os.path.join(temp_dir, os.path.basename(file))
        
        shutil.copy2(file, dst_file)
        stack = self.get_error_stack(dst_file, db_path, machine_name, mnbr, machine_type, last_marker)      #get error in log files

        if stack:
            #create query to add error to database
            data = self.create_query(stack) 
            sp_query = {
                "temp_table": "@",
                "data" : data
            }      
            sp_query_str = json.dumps(sp_query, ensure_ascii=False)
            # print(sp_query_str)
        #     #execute query
            self.enco_api_client.call_sp('TestDB..USP_PmMonitor_SetMachineLogs', {"@json_para": sp_query_str})

    #create query to insert data to db
    def create_query(self, stack): 
        result = []
        for machine_name, mnbr, datetime_event_str, id_error in stack:
            event_datetime_str = datetime_event_str.replace("/", "-")
            event_datetime = datetime.strptime(event_datetime_str, "%d-%m-%Y %H:%M:%S")
            datetime_str = event_datetime.strftime("%Y-%m-%d %H:%M:%S")

            try:
                datetime_event_local, datetime_event_utc = self.convert_time_utc(datetime_str)

                event_time = datetime_event_local.strftime("%Y-%m-%dT%H:%M:%S")     #event time in local timezone
                event_time_utc = datetime_event_utc.strftime("%Y-%m-%dT%H:%M:%S") + "Z"     #event time in utc timezone
                result.append({
                    "mnbr": mnbr,       #code machine
                    "event_time": event_time,       #event time in local timezone
                    "event_time_utc": event_time_utc,       #event time in utc timezone
                    "event_id": id_error
                })
                self.logger.info(f"Machine {machine_name}: Add OK: mnbr: {mnbr}, event_time:{event_time}, event_time_utc: {event_time_utc}, event_id: {id_error}")
            except Exception as e:
                self.logger.info(f"Machine {machine_name}: Add FAILED: mnbr: {mnbr}, event_time:{event_time}, event_time_utc: {event_time_utc}, event_id: {id_error}")
                continue
        return result
    
    #convert time in local tz to utc tz
    def convert_time_utc(self, event_time_str):
        try:
            dt_event = datetime.strptime(event_time_str, "%Y-%m-%d %H:%M:%S")
            event_local = dt_event.replace(tzinfo=timezone(timedelta(hours=7)))
            event_utc = event_local.astimezone(timezone.utc)
            return event_local, event_utc
        except Exception as e:
            self.logger.info(f"{e}")   

    # def set_last_marker_2(self, db_path, machine_name, last_marker):
    #         try:
    #             conn = sqlite3.connect(db_path, timeout=30)
    #             cursor = conn.cursor()      
    #             cursor.execute("""
    #                 UPDATE HEXA_machine_define 
    #                 SET last_marker = ?
    #                 WHERE machine_name = ?
    #             """, (last_marker, machine_name))

    #             conn.commit()
    #             conn.close() 
    #         except Exception as e:
    #             self.logger.error(f"Fail update last marker {machine_name}: {e}")   

    def run(self):
        self.logger.info(f"Starting {self.__class__.__name__} application")
                                       
        # add new event error to db
        # self.add_event_error()
        self.get_error_dict()

        self.load_error_dict()

        # #get HEXA_machine_define.db file containing local information
        db_path = os.path.join(self.local_dir, "HEXA_machine_define.db")
        copy_db_path = "HEXA_machine_define.db"
        if not os.path.exists(copy_db_path):
            shutil.copyfile(db_path, copy_db_path)
            self.create_last_marker(copy_db_path)       #create column last_marker

        # #get data in Qorvo_Machine_Define.db
        conn = sqlite3.connect(copy_db_path, timeout=30)
        cursor = conn.cursor()
        cursor.execute("SELECT machine_name, input_dir, pattern, mnbr, machine_type, last_marker FROM HEXA_machine_define")
        entries = cursor.fetchall()
        conn.close()

        # print(entries)
        # self.set_last_marker_2(copy_db_path, "TS-HT3#002", "2026-02-01 00:00:00")

        #run multithread
        def process_local_entry(entry):
                machine_name, input_dir, pattern, mnbr, machine_type, last_marker = entry
                last_marker = last_marker or ''
                self.parsing_ULT(copy_db_path, machine_name, input_dir, pattern, mnbr, machine_type, last_marker) 

        with ThreadPoolExecutor(max_workers=6) as executor:
            executor.map(process_local_entry, entries) 

        self.logger.info(f"{self.__class__.__name__} application finished")

